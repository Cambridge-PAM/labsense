import json
from pathlib import Path
from gsk_enviro_dict_temp import gsk_composite_red, gsk_composite_yellow, gsk_composite_green, gsk_inc_red, gsk_inc_yellow, gsk_inc_green, gsk_voc_red, gsk_voc_yellow, gsk_voc_green, gsk_aqua_red, gsk_aqua_yellow, gsk_aqua_green, gsk_air_red, gsk_air_yellow, gsk_air_green, gsk_health_red, gsk_health_yellow, gsk_health_green
import os
from dotenv import load_dotenv
import datetime

# Load environment variables from .env file in the repository root
repo_root = Path(__file__).resolve().parents[1]
env_path = repo_root / ".env"
load_dotenv(dotenv_path=env_path) 
import requests as req
import pandas as pd
from openpyxl import Workbook, load_workbook

from Labsense_SQL.constants import gsk_2016
# `gsk_2016` moved to `Labsense_SQL.constants` to avoid duplication.

chemical_dict = {
    "t-Butanol":"75-65-0",
    "IMS (Ethanol, Denatured)":"64-17-5",
    "Methanol":"67-56-1",
    "Tetrahydrofurfuryl Alcohol":"97-99-4",
    "2-Methoxyethanol":"109-86-4",
    "Glycerol Triacetate":"102-76-1",
    "Glycerol Diacetate":"111-55-7",
    "Isobutyl Acetate":"110-19-0",
    "Amyl Acetate":"628-63-7",
    "2-Ethylhexyl Acetate":"103-09-3",
    "Butyl Acetate":"123-86-4",
    "Methyl Oleate":"112-62-9",
    "Isoamyl Acetate":"123-92-2",
    "Isopropyl Acetate":"108-21-4",
    "Propyl Acetate":"109-60-4",
    "Dimethyl Succinate":"106-65-0",
    "n-Octyl Acetate":"112-14-1",
    "Ethyl Acetate":"141-78-6",
    "Ethyl Lactate":"97-64-3",
    "Diethyl Succinate":"123-25-1",
    "Dimethyl Adipate":"627-93-0",
    "gamma-Valerolactone":"108-29-2",
    "Diisopropyl Adipate":"6938-94-9",
    "Methyl Lactate":"547-64-8",
    "t-Butyl Acetate":"540-88-5",
    "Ethyl Formate":"109-94-4",
    "Methyl Acetate":"79-20-9",
    "Methyl Propionate":"554-12-1",
    "Ethyl Propionate":"105-37-3",
    "Methyl Formate":"107-31-3",
    "Propylene Carbonate":"108-32-7",
    "Ethylene Carbonate":"96-49-1",
    "Diethyl Carbonate":"105-58-8",
    "Dimethyl Carbonate":"616-38-6",
    "Butylene Carbonate":"4437-85-8",
    "Cyclopentanone":"120-92-3",
    "Cyclohexanone":"108-94-1",
    "3-Pentanone":"96-22-0",
    "Methylisobutyl Ketone":"108-10-1",
    "2-Pentanone":"107-87-9",
    "Methylethyl Ketone":"78-93-3",
    "Acetone":"67-64-1",
    "2,4,6-Collidine":"108-75-8",
    "Anisole":"100-66-3",
    "Ethoxybenzene":"103-73-1",
    "p-Xylene":"106-42-3",
    "Mesitylene":"108-67-8",
    "p-Cymene":"99-87-6",
    "Cumene":"98-82-8",
    "Toluene":"108-88-3",
    "Trifluorotoluene":"98-08-8",
    "Pyridine":"110-86-1",
    "Benzene":"71-43-2",
    "Isooctane":"540-84-1",
    "cis-Decalin":"493-01-6",
    "Heptane":"142-82-5",
    "L-Limonene":"5989-54-8",
    "Cyclohexane":"110-82-7",
    "D-Limonene":"5989-27-5",
    "Methylcyclohexane":"108-87-2",
    "Methylcyclopentane":"96-37-7",
    "Pentane":"109-66-0",
    "2-Methylpentane":"107-83-5",
    "Hexane":"110-54-3",
    "Petroleum Spirit":"8032-32-4",
    "Diethylene Glycol Monobutyl Ether":"112-34-5",
    "Dimethyl Isosorbide":"5306-85-4",
    "Dibutyl Ether":"142-96-1",
    "t-Amyl Methyl Ether":"994-05-8",
    "1,2,3-Trimethoxypropane":"20637-49-4",
    "Diphenyl Ether":"101-84-8",
    "t-Butyl Ethyl Ether":"637-92-3",
    "1,3-Dioxolane":"646-06-0",
    "Cyclopentyl Methyl Ether":"5614-37-9",
    "Diethoxymethane":"462-95-3",
    "2-Methyltetrahydrofuran":"96-47-9",
    "t-Butylmethyl Ether":"1634-04-4",
    "Diisopropyl Ether":"108-20-3",
    "Dimethoxymethane":"109-87-5",
    "Tetrahydrofuran":"109-99-9",
    "Bis(2-methoxyethyl)ether":"111-96-6",
    "1,4-Dioxane":"123-91-1",
    "Diethyl Ether":"60-29-7",
    "1,2-Dimethoxyethane":"110-71-4",
    "Dimethyl Ether":"115-10-6",
    "Dimethylpropylene Urea":"7226-23-5",
    "Dimethyl Sulphoxide":"67-68-5",
    "1,3-Dimethyl-2-imidazolidinone":"80-73-9",
    "Acetonitrile":"75-05-8",
    "Propanenitrile":"107-12-0",
    "Sulfolane":"126-33-0",
    "Formamide":"75-12-7",
    "N-Methyl Pyrrolidone":"872-50-4",
    "Dimethyl Acetamide":"127-19-5",
    "N-Ethylpyrrolidone":"2687-91-4",
    "N-Methylformamide":"123-39-7",
    "Dimethyl Formamide":"68-12-2",
    "Tetramethylurea":"632-22-4",
    "Carbon Disulfide":"75-15-0",
    "1,2,4-Trichlorobenzene":"120-82-1",
    "Chlorobenzene":"108-90-7",
    "1,2-Dichlorobenzene":"95-50-1",
    "Trichloroacetonitrile":"545-06-2",
    "Perfluorotoluene":"434-64-0",
    "Fluorobenzene":"462-06-6",
    "Perfluorocyclic Ether":"335-36-4",
    "Dichloromethane":"75-09-2",
    "1,2-Dichloroethane":"107-06-2",
    "Perfluorocyclohexane":"355-68-0",
    "Chloroform":"67-66-3",
    "Trichloroacetic Acid":"76-03-9",
    "Chloroacetic Acid":"79-11-8",
    "Trifluoroacetic Acid":"76-05-1",
    "Perfluorohexane":"355-42-0",
    "Carbon Tetrachloride":"56-23-5",
    "2,2,2-Trifluoroethanol":"75-89-8",
    "Furfural":"98-01-1",
    "N,N-Dimethylacetamide":"14433-76-2",
    "Dihydrolevoglucosenone":"1087696-49-8",
    "N,N-Dimethyloctanamide":"1118-92-9",
    "N,N-Dimethylaniline":"121-69-7",
    "Acetic Anhydride":"108-24-7",
    "Nitromethane":"75-52-5",
    "Triethylamine":"121-44-8",
    "Petroleum Ether":"64742-49-0",
    "Hexanes (Mixed Isomers)":"107-83-5",
}

#Dictionary to convert entries volume and mass to litres for summation
#Mass to litre conversion assumes an average density of 0.8 g/ml
from Labsense_SQL.constants import to_litre
# `to_litre` moved to `Labsense_SQL.constants` to avoid duplication.

composite_red_list=[]
composite_yellow_list=[]
composite_green_list=[]
inc_red_list=[]
inc_yellow_list=[]
inc_green_list=[]
voc_red_list=[]
voc_yellow_list=[]
voc_green_list=[]
aqua_red_list=[]
aqua_yellow_list=[]
aqua_green_list=[]
air_red_list=[]
air_yellow_list=[]
air_green_list=[]
health_red_list=[]
health_yellow_list=[]
health_green_list=[]

load_dotenv() #for getting the CHEMINVENTORY_CONNECTION_STRING, add the conncetion string in the .env file. If it doesn't exist, jus create one in the folder

def main():
    # Do ChemInventory processing
    for key, value in gsk_2016.items():
            ci = req.post("https://app.cheminventory.net/api/search/execute",
                        json = {"authtoken": os.getenv("CHEMINVENTORY_CONNECTION_STRING"),
                                "inventory": 873,
                                "type": "cas",
                                "contents": value})
            ci_json_raw = ci.json()
            if isinstance(ci_json_raw, str):
                ci_json_raw = json.loads(ci_json_raw)
            #print(ci_json_raw)

            # Normalize the JSON data
            ci_json_data = pd.json_normalize(ci_json_raw['data']['containers'])
            ci_df = pd.DataFrame(ci_json_data)

            if ci_df.empty:
                #print(f"No records for {key}")  # escape to allow for a null return
                temp_sum = 0
            else:
                # Remove any entries in "Missing - Stockcheck Only" location
                ci_df_real = ci_df.loc[ci_df["location"] != 527895]
                if ci_df_real.empty:
                    #print(f"No records for {key}")  # second escape if null return after filtering
                    temp_sum = 0
                else:
                    # Convert 'size' column to list and then to floats
                    size = list(ci_df_real['size'])
                    size_f = [float(i) for i in size]
                    unit = list(ci_df_real['unit'])
                    # Create blank list for litre-standardised conversion factor
                    conversion = []
                    # Assuming to_litre is a dictionary mapping units to conversion factors
                    for item in unit:
                        factor = to_litre.get(item)
                        conversion.append(factor)
                    # Calculate the total volume
                    temp = [size_f[i] * conversion[i] for i in range(len(size))]
                    temp_sum = sum(temp)
                    #print(f"Total volume for {key} is {temp_sum} litres -CAS {value}")
                    #adding volume into the right list
                    if value in gsk_composite_red.values():
                         composite_red_list.append(temp_sum)
                    if value in gsk_composite_yellow.values():
                         composite_yellow_list.append(temp_sum)
                    if value in gsk_composite_green.values():
                        composite_green_list.append(temp_sum)
                    if value in gsk_inc_red.values():
                         inc_red_list.append(temp_sum)
                    if value in gsk_inc_yellow.values():
                         inc_yellow_list.append(temp_sum)
                    if value in gsk_inc_green.values():
                        inc_green_list.append(temp_sum)
                    if value in gsk_voc_red.values():
                         voc_red_list.append(temp_sum)
                    if value in gsk_voc_yellow.values():
                         voc_yellow_list.append(temp_sum)
                    if value in gsk_voc_green.values():
                        voc_green_list.append(temp_sum)
                    if value in gsk_aqua_red.values():
                         aqua_red_list.append(temp_sum)
                    if value in gsk_aqua_yellow.values():
                         aqua_yellow_list.append(temp_sum)
                    if value in gsk_aqua_green.values():
                        aqua_green_list.append(temp_sum)
                    if value in gsk_air_red.values():
                         air_red_list.append(temp_sum)
                    if value in gsk_air_yellow.values():
                         air_yellow_list.append(temp_sum)
                    if value in gsk_air_green.values():
                        air_green_list.append(temp_sum)
                    if value in gsk_health_red.values():
                         health_red_list.append(temp_sum)
                    if value in gsk_health_yellow.values():
                         health_yellow_list.append(temp_sum)
                    if value in gsk_health_green.values():
                        health_green_list.append(temp_sum)

            
main()

def append_to_excel(file_path, sheet_name, new_rows_df):
    try:
        # Try to load the existing workbook
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            # If the sheet exists, read it into a DataFrame
            write_df = pd.read_excel(file_path, sheet_name=sheet_name)
            # Append the new rows
            write_df = pd.concat([write_df, new_rows_df], ignore_index=True)
        else:
            # If the sheet does not exist, use only the new rows
            write_df = new_rows_df
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        book = Workbook()
        write_df = new_rows_df
        book.save(file_path)  # Save the new workbook to create the file

    # Write back to the file, using append mode
    with pd.ExcelWriter(file_path, mode='a', engine="openpyxl", if_sheet_exists="replace") as writer:
        write_df.to_excel(writer, sheet_name=sheet_name, index=False)



sum_composite_red=sum(composite_red_list)
sum_composite_yellow=sum(composite_yellow_list)
sum_composite_green=sum(composite_green_list) #summing all the volumes for each colour


new_rows = [
    {"Date": datetime.datetime.now(), "Volume (L)": sum_composite_red, "Colour": "RED"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_composite_yellow, "Colour": "YELLOW"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_composite_green, "Colour": "GREEN"}
]
new_rows_df_composite = pd.DataFrame(new_rows)
append_to_excel('TrafficLightChemInventory.xlsx', 'Composite', new_rows_df_composite)


sum_inc_red=sum(inc_red_list)
sum_inc_yellow=sum(inc_yellow_list)
sum_inc_green=sum(inc_green_list) #summing all the volumes for each colour
new_rows = [
    {"Date": datetime.datetime.now(), "Volume (L)": sum_inc_red, "Colour": "RED"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_inc_yellow, "Colour": "YELLOW"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_inc_green, "Colour": "GREEN"}
]
new_rows_df_inc = pd.DataFrame(new_rows)
append_to_excel('TrafficLightChemInventory.xlsx', 'Incineration', new_rows_df_inc)


sum_voc_red=sum(voc_red_list)
sum_voc_yellow=sum(voc_yellow_list)
sum_voc_green=sum(voc_green_list) #summing all the volumes for each colour
new_rows = [
    {"Date": datetime.datetime.now(), "Volume (L)": sum_voc_red, "Colour": "RED"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_voc_yellow, "Colour": "YELLOW"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_voc_green, "Colour": "GREEN"}
]
new_rows_df_voc = pd.DataFrame(new_rows)
append_to_excel('TrafficLightChemInventory.xlsx', 'VOC emissions', new_rows_df_voc)



sum_aqua_red=sum(aqua_red_list)
sum_aqua_yellow=sum(aqua_yellow_list)
sum_aqua_green=sum(aqua_green_list)#summing all the volumes for each colour
new_rows = [
    {"Date": datetime.datetime.now(), "Volume (L)": sum_aqua_red, "Colour": "RED"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_aqua_yellow, "Colour": "YELLOW"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_aqua_green, "Colour": "GREEN"}
]
new_rows_df_aqua = pd.DataFrame(new_rows)
append_to_excel('TrafficLightChemInventory.xlsx', 'Aquatic Impact', new_rows_df_aqua)


sum_air_red=sum(air_red_list)
sum_air_yellow=sum(air_yellow_list)
sum_air_green=sum(air_green_list)#summing all the volumes for each colour
new_rows = [
    {"Date": datetime.datetime.now(), "Volume (L)": sum_air_red, "Colour": "RED"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_air_yellow, "Colour": "YELLOW"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_air_green, "Colour": "GREEN"}
]
new_rows_df_air = pd.DataFrame(new_rows)
append_to_excel('TrafficLightChemInventory.xlsx', 'Air Impact', new_rows_df_air)


sum_health_red=sum(health_red_list)
sum_health_yellow=sum(health_yellow_list)
sum_health_green=sum(health_green_list)#summing all the volumes for each colour
new_rows = [
    {"Date": datetime.datetime.now(), "Volume (L)": sum_health_red, "Colour": "RED"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_health_yellow, "Colour": "YELLOW"},
    {"Date": datetime.datetime.now(), "Volume (L)": sum_health_green, "Colour": "GREEN"}
]
new_rows_df_health = pd.DataFrame(new_rows)
append_to_excel('TrafficLightChemInventory.xlsx', 'Health Hazard', new_rows_df_health)
